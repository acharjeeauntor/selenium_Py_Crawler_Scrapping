
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
import os
from openpyxl import Workbook

baseUrl="https://www.scrapethissite.com/pages/simple/"
website_title = "Countries of the World: A Simple Example | Scrape This Site | A public sandbox for learning web scraping"
country_name_selector = ".country-name"
country_capital_name_selector = ".country-capital"
country_population_selector = ".country-population"

def browserOpen():
    chrome_options = Options()
    chrome_options.add_argument("--headless")
    chrome_options.add_argument("--disable-gpu")
    chrome_options.add_argument("--no-sandbox")
    chrome_options.add_argument("--start-maximized")

    driver = webdriver.Chrome(options=chrome_options)
    driver.get(baseUrl)
    driver.implicitly_wait(5)
    return driver

# Open the Browser with the URL
driver = browserOpen()
# verify the Web page load properly
assert website_title in driver.title
country_name_list = driver.find_elements(By.CSS_SELECTOR,country_name_selector)
capital_name_list = driver.find_elements(By.CSS_SELECTOR,country_capital_name_selector)
population_list = driver.find_elements(By.CSS_SELECTOR,country_population_selector)
total_country = len(country_name_list)

wb = Workbook()
ws = wb.active
ws['A1'] = 'Country Name'
ws['B1'] = 'Capital Name'
ws['C1'] = 'Population'
# Store Data into Excel File ---- Start From Here ----
for index in range(total_country):
    country_name = country_name_list[index].text
    capital_name = capital_name_list[index].text
    total_population = population_list[index].text
    ws.append([country_name,capital_name,total_population])

file_directory = 'Country_List_EXCEL_Data'
if not os.path.exists(file_directory):
    os.makedirs(file_directory)
        
file_path = os.path.join(file_directory, 'country_data.xlsx')

# Create the directory if it doesn't exist
wb.save(file_path)


# Close the Browser with the URL
driver.quit()


